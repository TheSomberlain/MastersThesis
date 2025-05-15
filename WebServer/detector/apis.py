import io
from django.http import FileResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import ProcessedImage
from .models import DetectedROI
from .models import Reports
from model_utils import process_with_yolo
from django.views.decorators.csrf import csrf_exempt
from WebServer import settings
import os
import uuid
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import csv
import shutil

UPLOAD_DIR = 'media/uploads'
os.makedirs(UPLOAD_DIR, exist_ok=True)

def temp_cleanup():
    imgs_to_delete = ProcessedImage.objects.filter(is_created_by_auth_user=False)
    for img in imgs_to_delete:
        rois = img.rois.all()
        for roi in rois:
            if roi.roi_file and os.path.isfile(roi.roi_file.path):
                os.remove(roi.roi_file.path)

        if img.image_file and os.path.isfile(img.image_file.path):
            os.remove(img.image_file.path)
        if img.original_file and os.path.isfile(img.original_file.path):
            os.remove(img.original_file.path)

    imgs_to_delete.delete()

@login_required
def report_images_api(request, report_id):
    try:
        report = Reports.objects.get(id=report_id, user=request.user)
        images_urls = [img.image_file.url for img in report.images.all()]
        return JsonResponse({'images': images_urls})
    except Reports.DoesNotExist:
        return JsonResponse({'images': []}, status=404)
    
@csrf_exempt
def upload_files_api(request):
    if request.method == 'POST':
        uploaded_files = request.FILES.getlist('files')
        saved_files = []
        for file in uploaded_files:
            file_path = os.path.join(UPLOAD_DIR, file.name)
            with open(file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            saved_files.append(file.name)
        return JsonResponse({'status': 'success', 'files': saved_files})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@csrf_exempt
def start_analysis_api(request):
    if request.method == 'POST':
        temp_cleanup()
        uploaded_files = request.FILES.getlist('files')
        if not uploaded_files:
            return JsonResponse({'status': 'error', 'message': 'No files provided.'}, status=400)
        is_authenticated = request.user.is_authenticated
        processed_images = []
        output_dir = os.path.join(settings.MEDIA_ROOT, 'results')
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(output_dir + "/rois", exist_ok=True)
        for uploaded_file in uploaded_files:
            try:
                temp_filename = f"{uuid.uuid4().hex}_{uploaded_file.name}"
                temp_path = os.path.join(output_dir, temp_filename)
                with open(temp_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)

                output_image, df, regions = process_with_yolo(temp_path)

                result_filename = f"result_{uuid.uuid4().hex}.png"
                result_path = os.path.join(output_dir, result_filename)
                output_image.save(result_path)

                processed_image = ProcessedImage.objects.create(image_file=f'results/{result_filename}',
                                                                original_file=uploaded_file.name,
                                                                is_created_by_auth_user=is_authenticated)

                for idx, roi in enumerate(regions):
                    roi_filename = f"rois/roi_{uuid.uuid4().hex}.png"
                    roi_path = os.path.join(output_dir, roi_filename)
                    roi['image'].save(roi_path)

                    DetectedROI.objects.create(
                        parent_image=processed_image,
                        roi_file=f'results/{roi_filename}',
                        label=roi['label'],
                        x_min=roi['xmin'],
                        y_min=roi['ymin'],
                        x_max=roi['xmax'],
                        y_max=roi['ymax']                        
                    )

                os.remove(temp_path)
                processed_images.append(processed_image)
            except Exception as e:
                print(f"[ERROR] Failed to process uploaded file: {e}")

        response_images = []
        for img in processed_images:
            response_images.append({
                'id': img.id,
                'image_url': img.image_file.url,
                'uploaded_at': img.uploaded_at.strftime('%Y-%m-%d %H:%M:%S'),
            })

        image_ids = [img.id for img in processed_images]
        request.session['last_image_ids'] = image_ids
        return JsonResponse({
            'status': 'completed',
            'redirect': '/result/'+ str(response_images[0]["id"]),
            'images_ids': image_ids,
            'is_authenticated': is_authenticated
        })
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@csrf_exempt
def export_rois_excel(request):
    image_ids_raw = request.POST.get('image_ids') or request.GET.get('image_ids')
    if image_ids_raw:
        image_ids = [int(i.strip()) for i in image_ids_raw.split(',')]
    else:
        image_ids = []

    images = ProcessedImage.objects.filter(id__in=image_ids)

    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    for img in images:
        rois = img.rois.all()
        sheet_name = img.original_file.name
        ws = wb.create_sheet(title=sheet_name)
        ws.append(['Image Name', 'ROI Image', 'Label', 'X min', 'Y min', 'X max', 'Y max'])

        for index, roi in enumerate(rois):
          
            roi_path = roi.roi_file.path
            image_name = img.original_file.name
            roi_path_1 = f"{roi.label}_{index + 1}" if roi.label else f"ROI_{index + 1}"
            roi_label = roi_path_1
            xl_image = XLImage(roi_path)
            row = ws.max_row + 1
            ws.row_dimensions[row].height = 100
            ws.column_dimensions['B'].width = 30

            ws.append([image_name, '', roi_label, roi.x_min, roi.y_min, roi.x_max, roi.y_max])

            cell_address = f'B{row}'
            xl_image.width = 100
            xl_image.height = 100
            ws.add_image(xl_image, cell_address)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return FileResponse(output, as_attachment=True, filename='rois_export.xlsx')   

@csrf_exempt
def export_rois_csv(request):
    try:
        image_ids_raw = request.POST.get('image_ids') or request.GET.get('image_ids')
        if image_ids_raw:
            image_ids = [int(i.strip()) for i in image_ids_raw.split(',')]
        else:
            image_ids = []

        images = ProcessedImage.objects.filter(id__in=image_ids)

        string_output  = io.StringIO()
        writer = csv.writer(string_output)
        writer.writerow(['Image Name', 'Label', 'X min', 'Y min', 'X max', 'Y max'])

        for img in images:
            image_name = img.original_file.name
            rois = img.rois.all()

            for index, roi in enumerate(rois):
                roi_label = f"{roi.label}_{index + 1}" if roi.label else f"ROI_{index + 1}"

                writer.writerow([
                    image_name,
                    roi_label,
                    roi.x_min,
                    roi.y_min,
                    roi.x_max,
                    roi.y_max
                ])

        bytes_output = io.BytesIO(string_output.getvalue().encode('utf-8'))
        bytes_output.seek(0)

        response = FileResponse(bytes_output, as_attachment=True, filename='rois_export.csv')
        response['Content-Type'] = 'text/csv'
        return response     
    except Exception as e:
            print("[ERROR] Export CSV failed:", e)
            return JsonResponse({'error': str(e)}, status=500)
    
@csrf_exempt
def create_report(request):
    try:
        image_ids_raw = request.POST.get('image_ids') or request.GET.get('image_ids')
        if image_ids_raw:
            image_ids = [int(i.strip()) for i in image_ids_raw.split(',')]
        else:
            image_ids = []

        images = ProcessedImage.objects.filter(id__in=image_ids)

        output_dir_excel = os.path.join(settings.MEDIA_ROOT, 'reports/xlsx/')
        output_dir_csv = os.path.join(settings.MEDIA_ROOT, 'reports/csv/')
        os.makedirs(output_dir_excel, exist_ok=True)
        os.makedirs(output_dir_csv, exist_ok=True)

        prefix = uuid.uuid4().hex
        excel_filename = f'report_{prefix}.xlsx'
        csv_filename = f'report_{prefix}.csv'
        output_file_path_excel = os.path.join(output_dir_excel, excel_filename)
        output_file_path_csv = os.path.join(output_dir_csv, csv_filename)

        wb = Workbook()

        if 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)

        for img in images:
            rois = img.rois.all()

            sheet_name = os.path.basename(img.original_file.name)
            sheet_name = sheet_name[:31]  
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['Image Name', 'ROI Image', 'Label', 'X min', 'Y min', 'X max', 'Y max'])

            for index, roi in enumerate(rois):
                roi_path = roi.roi_file.path
                image_name = os.path.basename(img.original_file.name)
                roi_label = f"{roi.label}_{index + 1}" if roi.label else f"ROI_{index + 1}"
                xl_image = XLImage(roi_path)
                row = ws.max_row + 1
                ws.row_dimensions[row].height = 100
                ws.column_dimensions['B'].width = 30

                ws.append([image_name, '', roi_label, roi.x_min, roi.y_min, roi.x_max, roi.y_max])

                cell_address = f'B{row}'
                xl_image.width = 100
                xl_image.height = 100
                ws.add_image(xl_image, cell_address)

        wb.save(output_file_path_excel)
        print(f"Report saved to {output_file_path_excel}")

        with open(output_file_path_csv, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Image Name', 'Label', 'X min', 'Y min', 'X max', 'Y max'])

            for img in images:
                image_name = os.path.basename(img.original_file.name)
                rois = img.rois.all()

                for index, roi in enumerate(rois):
                    roi_label = f"{roi.label}_{index + 1}" if roi.label else f"ROI_{index + 1}"
                    writer.writerow([
                        image_name,
                        roi_label,
                        roi.x_min,
                        roi.y_min,
                        roi.x_max,
                        roi.y_max
                    ])

        print(f"CSV report saved to {output_file_path_csv}")

        report = Reports.objects.create(
            user=request.user,
            result_csv=os.path.join('reports/csv/', csv_filename),
            result_xlsx=os.path.join('reports/xlsx/', excel_filename),
            name = f'report_{prefix}'
        )
        report.images.set(images)
        report.save()

        return JsonResponse({
            'report_id': report.id,
            'result_csv_url': f'/media/reports/csv/{csv_filename}',
            'result_xlsx_url': f'/media/reports/xlsx/{excel_filename}'
        })
    except Exception as e:
            print("[ERROR] Failed to create report: ", e)
            return JsonResponse({'error': str(e)}, status=500)